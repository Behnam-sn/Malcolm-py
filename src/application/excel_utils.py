from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

from config import Config


class Excel_Utils:
    @staticmethod
    def load_workbook(file_name: str) -> Workbook:
        return load_workbook(
            data_only=True,
            filename=file_name,
        )

    @staticmethod
    def create_workbook() -> Workbook:
        return Workbook()

    @staticmethod
    def create_sheet(workbook: Workbook, name: str) -> Worksheet:
        return workbook.create_sheet(name)

    @staticmethod
    def delete_sheet(workbook: Workbook, sheet: Worksheet) -> None:
        workbook.remove_sheet(sheet)

    @staticmethod
    def create_sheet_with_default_options(workbook: Workbook, name: str) -> Worksheet:
        sheet = Excel_Utils.create_sheet(workbook, name)
        Excel_Utils.change_sheet_direction_from_right_to_left(sheet)
        return sheet

    @staticmethod
    def delete_default_sheet(workbook: Workbook) -> None:
        sheet_names = workbook.sheetnames
        default_sheet_name = sheet_names[0]
        default_sheet = workbook.get_sheet_by_name(default_sheet_name)
        Excel_Utils.delete_sheet(workbook, default_sheet)

    @staticmethod
    def change_sheet_direction_from_right_to_left(sheet: Worksheet) -> None:
        sheet.sheet_view.rightToLeft = True

    @staticmethod
    def automatically_adjust_width_of_workbook_columns(workbook: Workbook) -> None:
        for sheet_name in workbook.sheetnames:
            sheet = workbook.get_sheet_by_name(sheet_name)
            Excel_Utils.automatically_adjust_width_of_sheet_columns(sheet)

    @staticmethod
    def automatically_adjust_width_of_sheet_columns(sheet) -> None:
        padding = 3

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + padding
            sheet.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def adjust_height_of_workbook_rows(workbook: Workbook, height: int) -> None:
        for sheet_name in workbook.sheetnames:
            sheet = workbook.get_sheet_by_name(sheet_name)
            Excel_Utils.adjust_height_of_sheet_rows(sheet, height)

    @staticmethod
    def adjust_height_of_sheet_rows(sheet, height: int) -> None:
        sheet_rows_length = sheet.max_row

        for row_id in range(1, sheet_rows_length + 1):
            sheet.row_dimensions[row_id].height = height

    @staticmethod
    def save_workbook(workbook: Workbook, path: str = ".", name: str = "excel") -> None:
        try:
            workbook.save(f"{path}\\{name}.xlsx")
        except PermissionError:
            print("Close Excel File Please")

    @staticmethod
    def append_list_of_string_to_sheet(
        sheet: Worksheet, items: list[str], cell_formatting_method
    ) -> None:
        row_index = Excel_Utils.compute_last_row_index(sheet)

        if row_index == 0:
            row_index = 1
        else:
            row_index += 1

        for column_index, value in enumerate(items):
            column_index += 1
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = value
            cell_formatting_method(cell)

    @staticmethod
    def append_dictionary_to_sheet(
        sheet: Worksheet,
        dictionary: dict,
        cell_formatting_method,
    ) -> None:
        row_index = Excel_Utils.compute_last_row_index(sheet)

        if row_index == 0:
            row_index = 1
        else:
            row_index += 1

        for column_index, value in enumerate(dictionary.values()):
            column_index += 1
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = value
            cell_formatting_method(cell)

    @staticmethod
    def append_list_of_dictionary_to_sheet(
        sheet: Worksheet, items: list[dict], cell_formatting_method
    ) -> None:
        row_index = Excel_Utils.compute_last_row_index(sheet)

        for item in items:
            row_index += 1

            for column_index, value in enumerate(item.values()):
                column_index += 1
                cell = sheet.cell(row=row_index, column=column_index)
                cell.value = value
                cell_formatting_method(cell)

    @staticmethod
    def compute_last_row_index(sheet: Worksheet) -> int:
        last_row_index = 0

        for row in sheet.iter_rows():
            last_row_index += 1

        return last_row_index

    @staticmethod
    def format_cell_default(cell: Cell) -> None:
        cell.font = Font(
            name=Config.FONT_NAME,
            size=Config.LINE_FONT_SIZE,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    @staticmethod
    def format_cell_headline(cell: Cell) -> None:
        cell.font = Font(
            name=Config.FONT_NAME,
            size=Config.HEADLINE_FONT_SIZE,
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    @staticmethod
    def format_cell_summary(cell: Cell) -> None:
        cell.font = Font(
            name=Config.FONT_NAME,
            size=Config.HEADLINE_FONT_SIZE,
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        color = Color(rgb=Config.SUMMARY_COLOR)
        cell.fill = PatternFill(patternType="solid", fgColor=color)

    @staticmethod
    def select_cell(sheet: Worksheet, row_index: int, column_index: int) -> Cell:
        return sheet.cell(row_index, column_index)

    @staticmethod
    def merge_cells(sheet: Worksheet, cell_1: Cell, cell_2: Cell) -> None:
        sheet.merge_cells(
            start_row=cell_1.row,
            start_column=cell_1.column,
            end_row=cell_2.row,
            end_column=cell_2.column,
        )

    @staticmethod
    def freeze_top_row(sheet: Worksheet):
        sheet.freeze_panes = "A2"
