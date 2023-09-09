from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

from application.excel_utils import Excel_Utils
from config import Config
from domain.processor import Processor
from domain.record import Record


class Utils:
    @staticmethod
    def convert_list_of_items_to_list_of_dictionaries(
        items: list,
    ) -> list[dict[str, str]]:
        list = []

        for item in items:
            list.append(item.convert_to_dictionary())

        return list

    @staticmethod
    def extract_headline_from_item(item: dict) -> list[str]:
        headlines = []

        for key in item:
            headlines.append(key)

        return headlines

    @staticmethod
    def extract_headline_from_items(items: list) -> list[str]:
        headlines = []
        item = items[0]

        for key in item:
            headlines.append(key)

        return headlines

    @staticmethod
    def generate_enter_and_exit_report_sheet(
        workbook: Workbook, records: list[Record]
    ) -> None:
        sheet = Excel_Utils.create_sheet_with_default_options(
            workbook=workbook,
            name=Config.ENTER_AND_EXIT_SHEET_NAME,
        )

        enter_and_exit = Processor.generate_enter_and_exit_items(records)
        items = Utils.convert_list_of_items_to_list_of_dictionaries(enter_and_exit)
        headlines = Utils.extract_headline_from_items(items)
        summary = Processor.generate_enter_and_exit_summary(
            enter_and_exit
        ).convert_to_dictionary()
        summary_headlines = Utils.extract_headline_from_item(summary)

        Excel_Utils.append_list_of_string_to_sheet(
            sheet=sheet,
            items=headlines,
            cell_formatting_method=Excel_Utils.format_cell_headline,
        )
        Excel_Utils.append_list_of_dictionary_to_sheet(
            sheet=sheet,
            items=items,
            cell_formatting_method=Excel_Utils.format_cell_default,
        )
        Excel_Utils.append_dictionary_to_sheet(
            sheet=sheet,
            dictionary=summary,
            cell_formatting_method=Excel_Utils.format_cell_summary,
        )
        Excel_Utils.append_list_of_string_to_sheet(
            sheet=sheet,
            items=summary_headlines,
            cell_formatting_method=Excel_Utils.format_cell_summary,
        )

        Utils.freeze_top_row(sheet=sheet)

    @staticmethod
    def generate_daily_report_sheet(workbook: Workbook, records: list[Record]) -> None:
        sheet = Excel_Utils.create_sheet_with_default_options(
            workbook=workbook,
            name=Config.DAILY_REPORT_SHEET_NAME,
        )

        items = Processor.generate_daily_items(records)
        dicts = Utils.convert_list_of_items_to_list_of_dictionaries(items)
        headlines = Utils.extract_headline_from_items(dicts)

        Excel_Utils.append_list_of_string_to_sheet(
            sheet=sheet,
            items=headlines,
            cell_formatting_method=Excel_Utils.format_cell_headline,
        )
        Excel_Utils.append_list_of_dictionary_to_sheet(
            sheet=sheet,
            items=dicts,
            cell_formatting_method=Excel_Utils.format_cell_default,
        )

        Utils.freeze_top_row(sheet=sheet)
        Utils.format_categories(sheet=sheet, categories_column_index=4)
        Utils.merge_identical_cells_by_column(sheet=sheet, column_index=1)
        Utils.merge_identical_cells_by_column(sheet=sheet, column_index=2)

    @staticmethod
    def generate_weekly_report_sheet(workbook: Workbook, records: list[Record]) -> None:
        sheet = Excel_Utils.create_sheet_with_default_options(
            workbook=workbook,
            name=Config.WEEKLY_REPORT_SHEET_NAME,
        )

        items = Processor.generate_weekly_items(records)
        dicts = Utils.convert_list_of_items_to_list_of_dictionaries(items)
        headlines = Utils.extract_headline_from_items(dicts)

        Excel_Utils.append_list_of_string_to_sheet(
            sheet=sheet,
            items=headlines,
            cell_formatting_method=Excel_Utils.format_cell_headline,
        )
        Excel_Utils.append_list_of_dictionary_to_sheet(
            sheet=sheet,
            items=dicts,
            cell_formatting_method=Excel_Utils.format_cell_default,
        )

        Utils.freeze_top_row(sheet=sheet)
        Utils.format_categories(sheet=sheet, categories_column_index=2)
        Utils.merge_identical_cells_by_column(sheet=sheet, column_index=1)
        Utils.merge_identical_cells_by_column(sheet=sheet, column_index=2)

    @staticmethod
    def generate_monthly_report_sheet(
        workbook: Workbook,
        records: list[Record],
    ) -> None:
        sheet = Excel_Utils.create_sheet_with_default_options(
            workbook=workbook,
            name=Config.MONTHLY_REPORT_SHEET_NAME,
        )

        monthly_items = Processor.generate_monthly_items(records)
        monthly_summary = Processor.generate_monthly_summary(monthly_items)
        dicts = monthly_summary.convert_to_list_of_dictionaries()

        # dicts = Utils.convert_list_of_items_to_list_of_dictionaries(monthly_items)
        # sorted_items = Processor.sort_items(items=dicts)
        headlines = Utils.extract_headline_from_items(dicts)

        summary = monthly_summary.convert_to_dictionary()

        Excel_Utils.append_list_of_string_to_sheet(
            sheet=sheet,
            items=headlines,
            cell_formatting_method=Excel_Utils.format_cell_headline,
        )
        Excel_Utils.append_list_of_dictionary_to_sheet(
            sheet=sheet,
            items=dicts,
            cell_formatting_method=Excel_Utils.format_cell_default,
        )
        Excel_Utils.append_dictionary_to_sheet(
            sheet=sheet,
            dictionary=summary,
            cell_formatting_method=Excel_Utils.format_cell_summary,
        )

        Utils.freeze_top_row(sheet=sheet)
        Utils.format_categories(sheet=sheet, categories_column_index=1)
        Utils.merge_identical_cells_by_column(sheet=sheet, column_index=1)
        Utils.merge_identical_cells_by_column(sheet=sheet, column_index=4)
        Utils.merge_identical_cells_by_column(sheet=sheet, column_index=5)

    @staticmethod
    def format_categories(sheet: Worksheet, categories_column_index: int) -> None:
        for row in sheet.iter_rows(
            min_col=categories_column_index, max_col=categories_column_index
        ):
            for cell in row:
                category = cell.value
                if category in Config.CATEGORIES_COLOR:
                    rgb_color = Config.CATEGORIES_COLOR[category]
                    color = Color(rgb=rgb_color)
                    cell.fill = PatternFill(patternType="solid", fgColor=color)

    @staticmethod
    def merge_identical_cells_by_column(sheet: Worksheet, column_index: int) -> None:
        last_row_index = Excel_Utils.compute_last_row_index(sheet)

        target_cell = None

        for row_index in range(1, last_row_index + 1):
            current_cell = Excel_Utils.select_cell(sheet, row_index, column_index)

            if row_index == 1:
                target_cell = current_cell
                continue

            if current_cell.value == target_cell.value and row_index != last_row_index:  # type: ignore
                continue

            cells_gape = current_cell.row - target_cell.row  # type: ignore
            if cells_gape == 1:
                target_cell = current_cell
                continue

            if row_index == last_row_index:
                Excel_Utils.merge_cells(sheet, target_cell, current_cell)  # type: ignore
                continue

            previous_cell = Excel_Utils.select_cell(sheet, row_index - 1, column_index)
            Excel_Utils.merge_cells(sheet, target_cell, previous_cell)  # type: ignore
            target_cell = current_cell

    @staticmethod
    def freeze_top_row(sheet: Worksheet):
        sheet.freeze_panes = "A2"
